"""
xlsx_builder.py — .xlsx workbook builder service.
Produces a formatted QA test suite workbook from structured test data.

Format:
  - Dashboard sheet: summary, stats, COUNTIF formulas
  - Per-section sheets: one per page/section with test cases
  - Dropdown validation on Status and Priority columns
  - Color-coded cells by status and priority
"""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint


# ── Color palette ──────────────────────────────────────────────────────────
COLORS = {
    # Header / branding
    "header_bg":    "1A1A2E",
    "header_fg":    "FFFFFF",
    "subheader_bg": "16213E",
    "accent":       "0F3460",
    "accent_light": "533483",

    # Status colors
    "pass":         "D4EDDA",
    "fail":         "F8D7DA",
    "blocked":      "FFF3CD",
    "not_run":      "E2E3E5",
    "in_progress":  "CCE5FF",

    # Priority colors
    "critical":     "FF4757",
    "high":         "FFA502",
    "medium":       "2ED573",
    "low":          "A4B0BE",

    # Misc
    "row_alt":      "F8F9FA",
    "border":       "DEE2E6",
    "white":        "FFFFFF",
    "light_gray":   "F1F3F5",
}

STATUS_OPTIONS = ["Not Run", "Pass", "Fail", "Blocked", "In Progress"]
PRIORITY_OPTIONS = ["Critical", "High", "Medium", "Low"]

STATUS_FILL = {
    "Pass":        PatternFill("solid", fgColor=COLORS["pass"]),
    "Fail":        PatternFill("solid", fgColor=COLORS["fail"]),
    "Blocked":     PatternFill("solid", fgColor=COLORS["blocked"]),
    "Not Run":     PatternFill("solid", fgColor=COLORS["not_run"]),
    "In Progress": PatternFill("solid", fgColor=COLORS["in_progress"]),
}

PRIORITY_FILL = {
    "Critical": PatternFill("solid", fgColor="FFD7DA"),
    "High":     PatternFill("solid", fgColor="FFEEBA"),
    "Medium":   PatternFill("solid", fgColor="D4EDDA"),
    "Low":      PatternFill("solid", fgColor="E2E3E5"),
}


def build_workbook(test_suite: dict) -> bytes:
    """
    Build the full .xlsx workbook from test_suite data.
    Returns raw bytes suitable for HTTP response streaming.
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default blank sheet

    sections = test_suite.get("sections", [])
    sheet_names = _make_sheet_names(sections)

    # Build per-section sheets first (Dashboard needs to reference them)
    for i, section in enumerate(sections):
        sheet_name = sheet_names[i]
        ws = wb.create_sheet(title=sheet_name)
        _build_section_sheet(ws, section, sheet_name)

    # Build Dashboard last (with COUNTIF references across sheets)
    ws_dash = wb.create_sheet(title="Dashboard", index=0)
    _build_dashboard(ws_dash, test_suite, sheet_names, sections)

    # Serialize to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Sheet builders ──────────────────────────────────────────────────────────

def _build_dashboard(ws, test_suite: dict, sheet_names: list, sections: list):
    """Build the Dashboard summary sheet."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3  # left margin

    # Title block
    ws.merge_cells("B2:H2")
    title_cell = ws["B2"]
    title_cell.value = f"QA Test Suite — {test_suite.get('site_name', 'Website')}"
    title_cell.font = Font(name="Calibri", size=20, bold=True, color=COLORS["header_fg"])
    title_cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 40

    # Metadata row
    ws.merge_cells("B3:H3")
    meta_cell = ws["B3"]
    meta_cell.value = (
        f"URL: {test_suite.get('base_url', '')}  |  "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
        f"Sections: {len(sections)}"
    )
    meta_cell.font = Font(name="Calibri", size=10, color="AAAAAA")
    meta_cell.fill = PatternFill("solid", fgColor=COLORS["subheader_bg"])
    meta_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 22

    # Summary text
    if test_suite.get("summary"):
        ws.merge_cells("B5:H6")
        summary_cell = ws["B5"]
        summary_cell.value = test_suite["summary"]
        summary_cell.font = Font(name="Calibri", size=11, color="333333", italic=True)
        summary_cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws.row_dimensions[5].height = 30
        ws.row_dimensions[6].height = 20

    # Stats header
    row = 8
    stat_headers = ["Section", "Sheet", "Total Tests", "Pass", "Fail", "Blocked", "Not Run", "In Progress"]
    stat_cols = ["B", "C", "D", "E", "F", "G", "H", "I"]

    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 12

    for col, hdr in zip(stat_cols, stat_headers):
        cell = ws[f"{col}{row}"]
        cell.value = hdr
        cell.font = Font(name="Calibri", size=10, bold=True, color=COLORS["header_fg"])
        cell.fill = PatternFill("solid", fgColor=COLORS["accent"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()
    ws.row_dimensions[row].height = 22

    # Stats rows
    total_tests = 0
    for i, (section, sheet_name) in enumerate(zip(sections, sheet_names)):
        row += 1
        tc_count = len(section.get("test_cases", []))
        total_tests += tc_count
        status_col = "H"  # Column H in each section sheet = Status

        fill = PatternFill("solid", fgColor=COLORS["row_alt"] if i % 2 else COLORS["white"])

        ws[f"B{row}"].value = section.get("name", sheet_name)
        ws[f"C{row}"].value = sheet_name
        ws[f"D{row}"].value = tc_count

        # COUNTIF formulas for each status
        for col, status in zip(["E", "F", "G", "H", "I"], STATUS_OPTIONS):
            cell = ws[f"{col}{row}"]
            cell.value = f'=COUNTIF(\'{sheet_name}\'!{status_col}:{status_col},"{status}")'
            cell.alignment = Alignment(horizontal="center")

        for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
            ws[f"{col}{row}"].fill = fill
            ws[f"{col}{row}"].border = _thin_border()
            ws[f"{col}{row}"].font = Font(name="Calibri", size=10)
        ws.row_dimensions[row].height = 18

    # Totals row
    row += 1
    ws[f"B{row}"].value = "TOTAL"
    ws[f"B{row}"].font = Font(name="Calibri", size=10, bold=True)
    ws[f"D{row}"].value = total_tests
    ws[f"D{row}"].font = Font(name="Calibri", size=10, bold=True)
    for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=COLORS["light_gray"])
        ws[f"{col}{row}"].border = _thin_border()

    # Sum formulas for totals
    first_data_row = 9
    last_data_row = row - 1
    for col in ["E", "F", "G", "H", "I"]:
        ws[f"{col}{row}"].value = f"=SUM({col}{first_data_row}:{col}{last_data_row})"
        ws[f"{col}{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"{col}{row}"].alignment = Alignment(horizontal="center")


def _build_section_sheet(ws, section: dict, sheet_name: str):
    """Build a single section sheet with all test cases."""
    ws.sheet_view.showGridLines = False

    # Column widths
    col_config = [
        ("A", 3),    # margin
        ("B", 10),   # ID
        ("C", 35),   # Title
        ("D", 45),   # Description
        ("E", 35),   # Preconditions
        ("F", 50),   # Steps
        ("G", 40),   # Expected Result
        ("H", 14),   # Status
        ("I", 12),   # Priority
        ("J", 15),   # Category
    ]
    for col, width in col_config:
        ws.column_dimensions[col].width = width

    # Header row
    ws.row_dimensions[1].height = 10  # top margin
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B2:J2")
    title_cell = ws["B2"]
    title_cell.value = section.get("name", sheet_name)
    title_cell.font = Font(name="Calibri", size=16, bold=True, color=COLORS["header_fg"])
    title_cell.fill = PatternFill("solid", fgColor=COLORS["header_bg"])
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Source URL
    if section.get("source_url"):
        ws.merge_cells("B3:J3")
        url_cell = ws["B3"]
        url_cell.value = f"Source: {section['source_url']}"
        url_cell.font = Font(name="Calibri", size=9, color="888888")
        url_cell.fill = PatternFill("solid", fgColor=COLORS["subheader_bg"])
        url_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[3].height = 16

    # Column headers
    headers = ["ID", "Test Case Title", "Description", "Preconditions", "Steps", "Expected Result", "Status", "Priority", "Category"]
    header_cols = ["B", "C", "D", "E", "F", "G", "H", "I", "J"]
    header_row = 5
    ws.row_dimensions[header_row].height = 22

    for col, hdr in zip(header_cols, headers):
        cell = ws[f"{col}{header_row}"]
        cell.value = hdr
        cell.font = Font(name="Calibri", size=10, bold=True, color=COLORS["header_fg"])
        cell.fill = PatternFill("solid", fgColor=COLORS["accent"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()

    # Data validations
    status_dv = DataValidation(
        type="list",
        formula1=f'"{",".join(STATUS_OPTIONS)}"',
        allow_blank=True,
        showDropDown=False,
    )
    priority_dv = DataValidation(
        type="list",
        formula1=f'"{",".join(PRIORITY_OPTIONS)}"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(status_dv)
    ws.add_data_validation(priority_dv)

    # Test case rows
    test_cases = section.get("test_cases", [])
    for i, tc in enumerate(test_cases):
        data_row = header_row + 1 + i
        ws.row_dimensions[data_row].height = 60

        row_fill = PatternFill("solid", fgColor=COLORS["row_alt"] if i % 2 else COLORS["white"])

        tc_id      = tc.get("id", f"TC-{i+1:03d}")
        title      = tc.get("title", "")
        desc       = tc.get("description", "")
        precon     = tc.get("preconditions", "")
        steps      = tc.get("steps", "").replace("\\n", "\n")
        expected   = tc.get("expected_result", "")
        status     = "Not Run"
        priority   = tc.get("priority", "Medium")
        category   = tc.get("category", "Functional")

        values = [tc_id, title, desc, precon, steps, expected, status, priority, category]

        for col, val in zip(header_cols, values):
            cell = ws[f"{col}{data_row}"]
            cell.value = val
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _thin_border()
            cell.fill = row_fill

        # Apply status color
        status_cell = ws[f"H{data_row}"]
        status_cell.fill = STATUS_FILL.get(status, row_fill)
        status_cell.alignment = Alignment(horizontal="center", vertical="center")
        status_dv.sqref += f"H{data_row}"

        # Apply priority color
        priority_cell = ws[f"I{data_row}"]
        priority_cell.fill = PRIORITY_FILL.get(priority, row_fill)
        priority_cell.alignment = Alignment(horizontal="center", vertical="center")
        priority_dv.sqref += f"I{data_row}"

        # Center ID and category
        ws[f"B{data_row}"].alignment = Alignment(horizontal="center", vertical="top")
        ws[f"J{data_row}"].alignment = Alignment(horizontal="center", vertical="top")

    # Freeze panes below header
    ws.freeze_panes = ws[f"B{header_row + 1}"]


# ── Helpers ─────────────────────────────────────────────────────────────────

def _thin_border() -> Border:
    side = Side(style="thin", color=COLORS["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def _make_sheet_names(sections: list) -> list:
    """Generate unique, Excel-safe sheet names (max 31 chars)."""
    names = []
    seen = {}
    for section in sections:
        raw = section.get("name", "Section")
        # Strip invalid Excel sheet name characters
        clean = "".join(c for c in raw if c not in r'\/:*?[]')
        clean = clean[:28].strip()
        if not clean:
            clean = "Section"
        # Deduplicate
        if clean in seen:
            seen[clean] += 1
            clean = f"{clean[:25]} {seen[clean]}"
        else:
            seen[clean] = 1
        names.append(clean)
    return names
